"""
批次爬取流程：
  1. 從 前百大公司.xlsx A 欄讀取 ticker，B 欄記錄狀態（True=已完成）
  2. 跳過 B 欄已為 True 的 ticker（斷點續跑）
  3. 先嘗試本地 JSON（不佔用 API 次數）
  4. 本地無資料才呼叫 API；若 API 回傳無季報 → 視為當日次數耗盡，立即停止
  5. 每筆成功上傳後即時將 B 欄寫為 True 並存檔

注意：爬蟲.py 不會被修改。
"""

import sys
import os
import json
import importlib.util

import pandas as pd
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# ── 載入 爬蟲.py ──────────────────────────────────────────────
_dir  = os.path.dirname(os.path.abspath(__file__))
_spec = importlib.util.spec_from_file_location(
    "scraper_module", os.path.join(_dir, "爬蟲.py")
)
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)

StatementDogScraper = _mod.StatementDogScraper
SESSION_COOKIE      = _mod.SESSION_COOKIE

import db_writer

EXCEL_PATH = os.path.join(_dir, "前百大公司.xlsx")
START_YEAR = 2016
END_YEAR   = 2026

# ── 工具函式 ──────────────────────────────────────────────────
def json_path(ticker):
    return os.path.join(_dir, f"{ticker}_raw.json")

def save_json(data, ticker):
    with open(json_path(ticker), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_json(ticker):
    p = json_path(ticker)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return None

def has_quarterly(data):
    return len(data.get("quarterly", {})) > 0

def write_status(wb, row, value):
    """將 B 欄第 row 列（1-based）寫入 True/False 並存檔。"""
    wb.active.cell(row=row, column=2, value=value)
    wb.save(EXCEL_PATH)

# ── 讀取 Excel ────────────────────────────────────────────────
df = pd.read_excel(EXCEL_PATH, header=0)
tickers     = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
# B 欄現有狀態（True 代表已完成）
done_flags  = df.iloc[:, 1].tolist() if df.shape[1] > 1 else [None] * len(tickers)

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
ws.cell(row=1, column=2, value="Status")   # 確保標頭存在

def _is_done(val) -> bool:
    if isinstance(val, bool):   return val
    if isinstance(val, (int, float)): return val == 1
    if isinstance(val, str):    return val.strip().lower() == "true"
    return False

pending = [(i, t) for i, (t, d) in enumerate(zip(tickers, done_flags), start=2)
           if not _is_done(d)]

print(f"共 {len(tickers)} 筆，已完成 {len(tickers)-len(pending)} 筆，"
      f"待處理 {len(pending)} 筆\n")

# ── 初始化爬蟲 ────────────────────────────────────────────────
scraper = StatementDogScraper(SESSION_COOKIE)

# ── 批次執行 ──────────────────────────────────────────────────
success = failed = 0

for row_idx, ticker in pending:
    print(f"[{row_idx-1:>3}/{len(tickers)}] {ticker}")

    data   = None
    source = ""

    # 優先讀本地 JSON（不耗用 API 次數）
    local = load_json(ticker)
    if local and has_quarterly(local):
        data   = local
        source = "本地 JSON"

    # 本地沒有才呼叫 API
    if data is None:
        try:
            fetched = scraper.fetch(ticker, START_YEAR, END_YEAR, delay=1.5)

            if has_quarterly(fetched):
                save_json(fetched, ticker)
                data   = fetched
                source = "API"
            else:
                # API 無季報 → 當日次數耗盡，停止
                print(f"  API 回傳無季報，當日次數已達上限，停止搜尋")
                print(f"  進度已儲存至 B 欄，明日從此繼續")
                wb.save(EXCEL_PATH)
                break

        except RuntimeError as e:
            print(f"  API 錯誤: {e}，略過")
            write_status(wb, row_idx, False)
            failed += 1
            continue
        except Exception as e:
            print(f"  例外: {e}，略過")
            write_status(wb, row_idx, False)
            failed += 1
            continue

    if data is None:
        print(f"  無資料，略過")
        write_status(wb, row_idx, False)
        continue

    # 上傳 DB
    print(f"  來源: {source}，季度指標: {len(data['quarterly'])} 個")
    try:
        db_writer.save(data, ticker)
        write_status(wb, row_idx, True)
        print(f"  ✓ 上傳成功，B 欄已更新")
        success += 1
    except Exception as e:
        print(f"  DB 寫入失敗: {e}")
        write_status(wb, row_idx, False)
        failed += 1

# ── 結果摘要 ──────────────────────────────────────────────────
print(f"\n{'='*50}")
print(f"本次：成功 {success} 筆 | 失敗 {failed} 筆")
already_done = sum(
    1 for d in done_flags
    if str(d).strip().upper() == "TRUE" or d is True
)

remaining = len(tickers) - already_done - success
print(f"Excel B 欄進度已儲存，剩餘約 {max(0, remaining)} 筆待明日繼續")
