"""
測試「前百大公司.xlsx」A 欄的每支股票是否能從 statementdog 爬取到資料，
結果寫入 B 欄（True / False）。
自含版本，不依賴 db_writer / supabase。
"""

import os
import sys
import time
import requests
import pandas as pd
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# ── Session Cookie（從 爬蟲.py 複製） ─────────────────────────
SESSION_COOKIE = "jD4yVGZj%2Bd9G6jTD4n88mbzPmOGmyWEQ2GtPiycYbuKfW0KA7mnX5I%2BY0SUTv7o6eStIpnW5FvTvYoNph2MWywNHNpgQnmJwFGQLWAiUkJH2TlkyyedPJql1EAM9OLWN%2BGg8hQwByxYQAGPSLMwvSdRHVfcdXi%2BAnsC%2B8fl33sp8qphoTm63SinU1g2D%2FlDAwwDY0X7HtHZ1lQ%2FJzEnVCixMQQm8UjI4dzUESa1isOOz1ZLP0t93Mpx1q%2FW6csYkD%2BKDzqIKyVWaouStnpylfT4Ypm5wopZu50Z%2Bihw4xFDLM%2Fj9dU4hEzaN%2Fn4JErQCG%2FwTYact1HxNHW48wVbTamAXC9S%2FgglhMxZDllMyHpCSmbg7LSr7EY1oXIoFvhLyphlIAv2RzBwygvpjLWyBn3ZX1cjtKy7QOi93jugtVZ9QnWtkfmbdHtSFvUd6yOaa%2Bu1wIZ5zy3oVOcjiSn97%2FzoH%2FbepWZQuL%2FNmvuArALSpm8zJ2qyAPSkHKPndDeBDhzoDXep6ksM9b3sp%2BAZIAEmwj%2FD56KQK8RT85%2FXTV36vQ1FbzqESOlovI%2B3QbEkW%2F%2B9nEwZl--1BMEDpyJ8mrTHzuv--q6nuYyQj62JGi9yBZnI%2FlQ%3D%3D"

BASE_URL = "https://statementdog.com"
API_BASE = "https://statementdog.com/api/v2/fundamentals"

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "前百大公司.xlsx")

# ── 建立 Session ───────────────────────────────────────────────
session = requests.Session()
session.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
})
session.cookies.set("_statementdog_session_v2", SESSION_COOKIE, domain="statementdog.com")

resp = session.get(f"{BASE_URL}/")
if '"is_signed_in":true' in resp.text:
    print("[初始化] 登入成功 ✓")
else:
    print("[初始化] ⚠️  Cookie 無效或已過期，請重新從 F12 複製")

# ── 讀取 A 欄 Ticker（跳過第 1 列標頭） ─────────────────────────
df = pd.read_excel(EXCEL_PATH, header=0)
tickers = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
print(f"共 {len(tickers)} 支股票待測試\n")

# ── 逐一測試 ──────────────────────────────────────────────────
results = []
for i, ticker in enumerate(tickers, 1):
    ok = False
    try:
        analysis_url = f"{BASE_URL}/analysis/{ticker}"
        session.get(analysis_url)
        time.sleep(1.0)

        url = f"{API_BASE}/{ticker}/2023/2026/cf?qbu=true&qf=analysis"
        resp = session.get(url, headers={"Referer": analysis_url}, timeout=15)
        data = resp.json()

        if "error" not in data:
            pe_data = data.get("monthly", {}).get("PE", {}).get("data", [])
            non_null = [v for _, v in pe_data if v is not None]
            ok = len(non_null) >= 1
    except Exception as e:
        print(f"  [{ticker}] 例外: {e}")

    results.append(ok)
    status = "✓ True " if ok else "✗ False"
    print(f"[{i:>3}/{len(tickers)}] {ticker:<12} → {status}")

# ── 將結果寫回 Excel B 欄（從第 2 列開始，第 1 列寫標頭） ──────
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
ws.cell(row=1, column=2, value="Can_Scrape")

for row_idx, value in enumerate(results, start=2):
    ws.cell(row=row_idx, column=2, value=value)

wb.save(EXCEL_PATH)
print(f"\n完成！結果已寫入 B 欄：{os.path.abspath(EXCEL_PATH)}")
print(f"可爬取：{sum(results)} / {len(results)}")
